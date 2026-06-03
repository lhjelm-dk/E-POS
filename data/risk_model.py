"""Risk Model — structure, I/O and runtime conversion.

A RiskModel defines the petroleum system element hierarchy:

  Pillar (e.g. Charge)
    ├── Play element     (single assessment at play scale, or sub-groups)
    └── Conditional      (one or more groups of leaf elements)
          Group (e.g. Migration)
            └── Element  (leaf: carries S_for, S_against, success_criteria …)

Assessment values (S_for, S_against, evidence text) are stored separately in an
*assessment_data* dict keyed by node_id, merged at runtime.  This separation lets
the same template be applied to different prospects without losing data.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

MODELS_DIR = Path(__file__).parent / "models"

# ── Operator vocabulary ─────────────────────────────────────────────────────────
# Canonical definitions live in logic/esl_pipeline.py (the one module that uses
# them at runtime).  Re-exported here so downstream callers that import from
# data.risk_model don't need a separate import from logic.esl_pipeline.

from logic.esl_pipeline import (  # noqa: E402 — module-level import after constants
    ESL_MODE_OPTIONS as _ESL_PIPELINE_MODE_OPTIONS,
    CLASSIC_POS_OPERATOR_OPTIONS,
    DEFAULT_CLASSIC_POS_MODE as DEFAULT_CLASSIC_POS_OPERATOR,
    ESL_TO_CLASSIC_RECOMMENDATION,
)
from logic.session_keys import SK  # noqa: E402

ESL_OPERATOR_OPTIONS: list[str] = list(_ESL_PIPELINE_MODE_OPTIONS)

ESL_OPERATOR_DOCS: dict[str, str] = {
    "ESL-ALL (min/min)": (
        "**Weakest link.** `POS = min(all element POS)`. "
        "Use when ALL conditions must hold simultaneously. "
        "A single failing element kills the group / pillar.  "
        "*References: Rose 2001; Milkov 2015.*"
    ),
    "ESL-ANY (max/max)": (
        "**Best-case / disjunctive.** `POS = max(all element POS)`. "
        "Use when ANY one of several alternatives suffices "
        "(e.g. either structural OR stratigraphic closure).  "
        "*ESL equivalent of OR-logic.*"
    ),
    "ESL-IPT (sufficiency/dependency)": (
        "**Independence-adjusted (sufficiency).** "
        "Dependency ρ ∈ [0, 1]: 0 = fully independent, 1 = fully correlated. "
        "POS is computed with combined ESL masses using the IPT algorithm.  "
        "Use when evidence is partially shared between elements.  "
        "*Based on Dezert–Smarandache / Dempster–Shafer combination.*"
    ),
    "Product (Π)": (
        "**Multiplicative.** `POS = Π(all element POS)`. "
        "Each element is treated as independent; failure of any one is fatal. "
        "Equivalent to the classic Rose / AAPG four-factor product.  "
        "Note: applies product to both S_for and S_against masses."
    ),
    "Mean": (
        "**Arithmetic average.** `POS = mean(all element POS)`. "
        "Only appropriate when alternatives represent equal-weight, "
        "non-exclusive lines of evidence.  "
        "⚠️ Not coherent under ESL theory — use cautiously."
    ),
}

# Excel aggregation_rule → model operator
_XLSX_RULE_TO_ESL: dict[str, str] = {
    "min":     "ESL-ALL (min/min)",
    "max":     "ESL-ANY (max/max)",
    "product": "Product (Π)",
    "mean":    "Mean",
    "esl_pos": "ESL-ALL (min/min)",  # leaf node — rule is irrelevant at leaf level
}

PILLAR_COLORS_DEFAULT: dict[str, str] = {
    "Charge":    "#F69292",
    "Closure":   "#8CB7FC",
    "Reservoir": "#FFD44B",
    "Retention": "#B5E6A2",
}

# Rich guidance text for standard play pillars — preserved from the original build_models()
# Single source of truth for play-pillar description / considerations guidance.
# Consumed by app.build_models() (the hardcoded fallback default prospect),
# components.model_builder (new-pillar prefill), and build_runtime_models()
# below.  Keeping one copy prevents the live-prospect text and the model-builder
# text from drifting apart.
_PLAY_GUIDANCE: dict[str, dict[str, str]] = {
    "Charge": {
        "description": (
            "PLAY-LEVEL: The probability that the shared charge elements — source richness, "
            "maturation, expulsion, and regional migration pathway — are sufficient to charge "
            "at least one accumulation anywhere in the play area. "
            "DO NOT include prospect-specific migration path — that belongs in Conditional Charge."
        ),
        "considerations": (
            "Things to consider when assigning play charge risk:\n\n"
            "• Lateral migration: Is there evidence of carrier bed continuity and migration fairway geometry "
            "that would allow hydrocarbons to reach this play area from the kitchen? Consider lateral seal "
            "integrity along the migration pathway and whether thief sands or faults could divert charge.\n\n"
            "• Source rock presence: Is a source rock proven or inferred in the play area or kitchen? "
            "Consider organic richness (TOC), type (kerogen type I/II/III), and whether it is in the "
            "stratigraphic section expected to be mature.\n\n"
            "• HC generation and maturation: Is the source rock at the right maturity window for the "
            "target phase (oil window ~0.6–1.3% Ro; gas condensate ~1.3–2.0% Ro)? Consider burial history "
            "and heat flow calibration.\n\n"
            "• Timing of migration: Has migration occurred, or is it ongoing? This timing consideration "
            "should be embedded in your play assessment — see timing notes embedded in Closure and "
            "Retention below.\n\n"
            "• Column height adequacy: Is the expected fetch area and source volume sufficient to fill "
            "the closure to at least P90 EUR? Consider drainage area, expulsion efficiency, and migration losses.\n\n"
            "• Fluid phase prediction: What HC phase is expected (dry gas, wet gas, condensate, oil)? "
            "Fluid phase prediction affects commerciality, volumetrics, and seal requirements. "
            "Consider GOR from geochemical modeling, API gravity from analogues, and whether the phase "
            "matches the commercial threshold for P90 EUR."
        ),
    },
    "Closure": {
        "description": (
            "The probability that a valid trapping geometry (closure) is present at the correct stratigraphic level "
            "to contain at least one accumulation in the play area with minimum hydrocarbon volume or more. "
            "This pillar assesses GEOMETRY ONLY — 4-way dip closure, 3-way fault-assisted, stratigraphic pinch-out, "
            "angular truncation, or combined structural-stratigraphic traps. "
            "Seal capacity is NOT assessed here — it is assessed under Retention."
        ),
        "considerations": (
            "Things to consider when assigning play closure risk:\n\n"
            "• Closure type: What is the trapping mechanism? Is it structural (4-way, 3-way), stratigraphic "
            "(pinch-out, channel belt truncation), or combined?\n\n"
            "• Seismic confidence: What is the structural confidence on the mapped closure? "
            "Is the closure defined on depth-converted 3D seismic, or regional 2D with large velocity uncertainty?\n\n"
            "• Depth conversion uncertainty: Could velocity errors open or close the structure? "
            "Consider the P10/P90 spill-point depth range.\n\n"
            "• Do NOT assess seal here. Seal capacity and integrity belong in the Retention pillar."
        ),
    },
    "Reservoir": {
        "description": (
            "The probability that an effective reservoir of sufficient quality is present "
            "to contain at least one accumulation in the play area with minimum hydrocarbon "
            "volume or more"
        ),
        "considerations": (
            "Things to consider when assigning play reservoir risk:\n\n"
            "• Reservoir rock presence: Is a reservoir-quality rock (sandstone, carbonate, fractured basement) "
            "proven or predicted at the target level? Consider depositional environment and diagenetic history.\n\n"
            "• Reservoir effectiveness: Does the reservoir have sufficient porosity and permeability to deliver "
            "commercial flow rates? Consider net-to-gross, facies variability, and diagenetic overprinting.\n\n"
            "• Reservoir heterogeneity and compartmentalization: Consider whether baffles, tight streaks, or "
            "stratigraphic pinch-outs could compartmentalize the reservoir and reduce effective drainage volume. "
            "Highly heterogeneous reservoirs may have a larger EUR uncertainty range. This is NOT a separate "
            "risk element here but should be reflected in your volumetric P10/P90 spread.\n\n"
            "• Drive mechanism: What is the primary drive? Aquifer support, gas cap expansion, solution gas drive, "
            "or compaction drive? This affects recovery factor and should be considered in EUR P90 calibration.\n\n"
            "• HC column height adequacy: Is the expected HC column height sufficient to meet P90 EUR? "
            "This is the bridge between Closure (geometry) and Reservoir (deliverability). A large closure with "
            "a thin net pay column may not meet the minimum commercial threshold."
        ),
    },
    "Retention": {
        "description": (
            "The probability that hydrocarbons are contained by seals and preserved from "
            "spillage and degradation in at least one accumulation in the play area with "
            "minimum hydrocarbon volume or more"
        ),
        "considerations": (
            "Things to consider when assigning play retention risk:\n\n"
            "• Top seal capacity: Lithological seal (shale, evaporite, tight carbonate) — is it regionally continuous "
            "and of sufficient capillary entry pressure to hold the expected HC column?\n\n"
            "• Lateral seal and fault seal: For fault-bounded traps, what is the confidence in across-fault seal? "
            "Consider shale gouge ratio (SGR), juxtaposition, and diagenetic cementation.\n\n"
            "• Base seal: Is there a basal seal preventing downward migration losses (relevant where underlying "
            "reservoir-quality units are present)?\n\n"
            "• Preservation from spillage: Has the structural geometry been stable? Consider: (a) tilt-and-spill "
            "from later tilting events; (b) seal breaching from fault reactivation post-charge; "
            "(c) flushing by meteoric water influx or pressure depletion.\n\n"
            "• Preservation from degradation:\n"
            "  - Biodegradation (bacterial): The probability that the accumulation is NOT biodegraded by anaerobic "
            "bacteria to a degree that prevents production of P90 recoverable oil. Risk is elevated at reservoir "
            "temperatures <80°C and where meteoric water recharge pathways exist.\n"
            "  - Thermal cracking (over-maturation): The probability that liquid hydrocarbons are NOT cracked to "
            "dry gas or pyrobitumen by excessive burial temperatures (>150–160°C for oil). Consider burial history "
            "and present-day reservoir temperature.\n"
            "  - Oxidation: The probability of no oxidative degradation along shallow migration pathways.\n\n"
            "⚠️ NOTE: Seal capacity belongs here (Retention), NOT in Closure. The Closure pillar assesses geometry only."
        ),
    },
}


# ── Dataclasses ─────────────────────────────────────────────────────────────────

@dataclass
class RiskElement:
    """A leaf element in the risk model (contributes directly to S_for/S_against)."""
    node_id: str
    name: str
    success_criteria: str = ""
    description: str = ""
    considerations: str = ""
    default_s_n: float = 0.5
    default_s_neg: float = 0.1
    default_w: float = 0.5
    notes: str = ""


@dataclass
class ElementGroup:
    """A named group of leaf elements within a conditional pillar."""
    group_id: str
    group_label: str
    aggregation_rule: str = "ESL-ALL (min/min)"
    classic_pos_aggregation_rule: str = "Min (weakest link)"
    elements: list[RiskElement] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.elements = [
            RiskElement(**e) if isinstance(e, dict) else e
            for e in self.elements
        ]


@dataclass
class PlayElement:
    """Play-level element (single assessment or set of sub-groups)."""
    node_id: str
    name: str
    success_criteria: str = ""
    description: str = ""
    considerations: str = ""
    default_s_n: float = 0.5
    default_s_neg: float = 0.1
    default_w: float = 0.5
    aggregation_rule: str = "ESL-ALL (min/min)"
    sub_groups: list[ElementGroup] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.sub_groups = [
            ElementGroup(**g) if isinstance(g, dict) else g
            for g in self.sub_groups
        ]


@dataclass
class PillarDef:
    """One pillar (e.g. Charge) with its play element and conditional groups."""
    pillar_id: str          # internal key used in models dict
    display_name: str       # shown in the UI
    color: str = "#e5e7eb"
    play: PlayElement = field(default_factory=lambda: PlayElement(node_id="", name=""))
    conditional_operator: str = "ESL-ALL (min/min)"
    classic_pos_operator: str = "Min (weakest link)"
    conditional_groups: list[ElementGroup] = field(default_factory=list)

    def __post_init__(self) -> None:
        if isinstance(self.play, dict):
            self.play = PlayElement(**self.play)
        self.conditional_groups = [
            ElementGroup(**g) if isinstance(g, dict) else g
            for g in self.conditional_groups
        ]


@dataclass
class RiskModel:
    """Top-level risk model template — defines element structure, not assessment values."""
    model_id: str
    model_name: str
    version: str = "1.0"
    description: str = ""
    pillars: list[PillarDef] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.pillars = [
            PillarDef(**p) if isinstance(p, dict) else p
            for p in self.pillars
        ]

    # ── Serialization ──────────────────────────────────────────────────────────

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: dict) -> "RiskModel":
        return cls(**d)

    def save(self, path: "Path | str | None" = None) -> Path:
        MODELS_DIR.mkdir(parents=True, exist_ok=True)
        if path is None:
            path = MODELS_DIR / f"{self.model_id}.json"
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(self.to_dict(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path

    @classmethod
    def load(cls, path: "Path | str") -> "RiskModel":
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        return cls.from_dict(data)

    # ── Helpers ────────────────────────────────────────────────────────────────

    def all_leaf_node_ids(self) -> list[str]:
        ids: list[str] = []
        for p in self.pillars:
            pe = p.play
            if not pe.sub_groups:
                if pe.node_id:
                    ids.append(pe.node_id)
            else:
                for g in pe.sub_groups:
                    for e in g.elements:
                        ids.append(e.node_id)
            for g in p.conditional_groups:
                for e in g.elements:
                    ids.append(e.node_id)
        return ids

    def find_play_def(self, pillar_id: str) -> "PillarDef | None":
        return next((p for p in self.pillars if p.pillar_id == pillar_id), None)


# ── Model listing ─────────────────────────────────────────────────────────────

def list_saved_models() -> list[Path]:
    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(MODELS_DIR.glob("*.json"))


# ── Excel import ──────────────────────────────────────────────────────────────

def from_xlsx(
    path: "Path | str",
    model_id: str = "imported",
    model_name: str = "Imported Model",
) -> RiskModel:
    """Parse a risk_model_nodes.xlsx (or compatible) file into a RiskModel.

    Expected sheet name: 'nodes'.
    Required columns: node_id, parent_id, name, branch, pillar_group_id,
                      aggregation_rule, s_n, s_neg_n, w, success_criteria, notes.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel import: pip install openpyxl")

    wb = openpyxl.load_workbook(str(path), data_only=True)
    if "nodes" not in wb.sheetnames:
        raise ValueError("Excel file must contain a sheet named 'nodes'.")
    ws = wb["nodes"]

    headers: list[str] = []
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if not headers:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))

    play_rows = [r for r in rows if str(r.get("branch", "")).strip() == "play"]
    cond_pillar_rows = [
        r for r in rows
        if str(r.get("branch", "")).strip() == "conditional"
        and str(r.get("parent_id", "")).strip() == "ROOT"
    ]

    pillars: list[PillarDef] = []

    for pp_row in sorted(play_rows, key=lambda r: r.get("pillar_group_id") or 0):
        pg_id = pp_row.get("pillar_group_id")
        cp_row = next(
            (r for r in cond_pillar_rows if r.get("pillar_group_id") == pg_id),
            None,
        )

        raw_name = str(pp_row.get("name") or "").replace("(play)", "").replace("(Play)", "").strip()
        # Internal pillar_id "Closure" is the canonical name. Old user files that say "Trap"
        # are mapped to "Closure" here (one-way; no reverse migration).
        pillar_id = "Closure" if raw_name.lower() in ("closure", "trap") else raw_name
        color = PILLAR_COLORS_DEFAULT.get(pillar_id, PILLAR_COLORS_DEFAULT.get(raw_name, "#e5e7eb"))

        guidance = _PLAY_GUIDANCE.get(pillar_id, {})

        play_nid = str(pp_row.get("node_id") or f"{pillar_id.upper()}_PLAY")
        play_elem = PlayElement(
            node_id=play_nid,
            name=raw_name,
            success_criteria=str(pp_row.get("success_criteria") or ""),
            description=guidance.get("description", str(pp_row.get("notes") or "")),
            considerations=guidance.get("considerations", ""),
            default_s_n=_float_safe(pp_row.get("s_n"), 0.5),
            default_s_neg=_float_safe(pp_row.get("s_neg_n"), 0.1),
            default_w=_float_safe(pp_row.get("w"), 0.5),
        )

        # Conditional groups from children of the conditional pillar node
        cond_groups: list[ElementGroup] = []
        cond_op = "ESL-ALL (min/min)"
        if cp_row:
            cp_id = str(cp_row.get("node_id") or "")
            cond_op = _XLSX_RULE_TO_ESL.get(
                str(cp_row.get("aggregation_rule") or "min"), "ESL-ALL (min/min)"
            )
            grp_rows = [r for r in rows if str(r.get("parent_id") or "") == cp_id]
            for g_row in grp_rows:
                g_id = str(g_row.get("node_id") or "")
                g_label = str(g_row.get("name") or "Group").strip()
                g_op = _XLSX_RULE_TO_ESL.get(
                    str(g_row.get("aggregation_rule") or "min"), "ESL-ALL (min/min)"
                )
                leaf_rows = [r for r in rows if str(r.get("parent_id") or "") == g_id]
                elements = [
                    RiskElement(
                        node_id=str(lr.get("node_id") or ""),
                        name=str(lr.get("name") or ""),
                        success_criteria=str(lr.get("success_criteria") or ""),
                        description=str(lr.get("notes") or ""),
                        considerations="",
                        default_s_n=_float_safe(lr.get("s_n"), 0.5),
                        default_s_neg=_float_safe(lr.get("s_neg_n"), 0.1),
                        default_w=_float_safe(lr.get("w"), 0.5),
                        notes=str(lr.get("notes") or ""),
                    )
                    for lr in leaf_rows
                ]
                if elements:
                    cond_groups.append(
                        ElementGroup(
                            group_id=g_id,
                            group_label=g_label,
                            aggregation_rule=g_op,
                            elements=elements,
                        )
                    )

        pillars.append(
            PillarDef(
                pillar_id=pillar_id,
                display_name=raw_name,
                color=color,
                play=play_elem,
                conditional_operator=cond_op,
                conditional_groups=cond_groups,
            )
        )

    return RiskModel(
        model_id=model_id,
        model_name=model_name,
        version="1.0",
        description="Imported from risk_model_nodes.xlsx",
        pillars=pillars,
    )


# ── Excel export ──────────────────────────────────────────────────────────────

def to_xlsx(model: RiskModel, path: "Path | str | Any") -> "Path | None":
    """Export a RiskModel to risk_model_nodes.xlsx format.

    path may be a file-system path (str / Path) or any file-like object
    accepted by openpyxl.Workbook.save() (e.g. io.BytesIO).
    Returns the Path if a filesystem path was given, else None.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    _is_path = isinstance(path, (str, Path))
    if _is_path:
        path = Path(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "nodes"

    HDR = [
        "node_id", "parent_id", "name", "branch", "pillar_group_id",
        "aggregation_rule", "s_n", "s_neg_n", "w", "success_criteria", "notes",
    ]
    ws.append(HDR)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1e3a5f")

    _esl_to_xlsx: dict[str, str] = {
        "ESL-ALL (min/min)":              "min",
        "ESL-ANY (max/max)":              "max",
        "ESL-IPT (sufficiency/dependency)": "min",
        "Product (Π)":                    "product",
        "Mean":                           "mean",
    }

    ws.append(["ROOT", "", "Final GCOS", "root", "", "product", "", "", "", "", ""])

    for pg_id, pdef in enumerate(model.pillars, 1):
        pid = pdef.pillar_id
        dn = pdef.display_name
        pe = pdef.play
        play_nid = pe.node_id or f"{pid.upper()}_PLAY"

        ws.append([
            play_nid, "ROOT", f"{dn} (play)", "play", pg_id, "esl_pos",
            pe.default_s_n, pe.default_s_neg, pe.default_w,
            pe.success_criteria, pe.description,
        ])

        cp_id = f"{pid.upper()}_COND"
        ws.append([
            cp_id, "ROOT", f"{dn} (conditional)", "conditional", pg_id,
            _esl_to_xlsx.get(pdef.conditional_operator, "min"),
            "", "", "", "", "",
        ])

        for grp in pdef.conditional_groups:
            g_id = grp.group_id or _slug(f"{cp_id}_{grp.group_label}")
            ws.append([
                g_id, cp_id, grp.group_label, "conditional", pg_id,
                _esl_to_xlsx.get(grp.aggregation_rule, "min"),
                "", "", "", "", "",
            ])
            for elem in grp.elements:
                ws.append([
                    elem.node_id, g_id, elem.name, "conditional", pg_id, "esl_pos",
                    elem.default_s_n, elem.default_s_neg, elem.default_w,
                    elem.success_criteria, elem.notes or elem.description,
                ])

    # Legend sheet
    ws2 = wb.create_sheet("legend")
    for row in [
        ("column", "type", "description", "example"),
        ("node_id", "str", "Unique ID. Underscores, no spaces.", "CHARGE_PLAY"),
        ("parent_id", "str", "node_id of parent. Empty for ROOT only.", "ROOT"),
        ("name", "str", "Display name.", "Charge (play)"),
        ("branch", "str", '"root" | "play" | "conditional"', "conditional"),
        ("pillar_group_id", "int", "Links play pillar to conditional counterpart.", "1"),
        ("aggregation_rule", "str", '"product" | "min" | "max" | "mean" | "esl_pos"', "min"),
        ("s_n", "float", "Default ESL S_for [0–1]. Leaf nodes only.", "0.55"),
        ("s_neg_n", "float", "Default ESL S_against [0–1]. Leaf nodes only.", "0.15"),
        ("w", "float", "Default stance [0–1].", "0.5"),
        ("success_criteria", "str", "What must be true for success.", "Sufficient to fill P99 EUR"),
        ("notes", "str", "Free-text description / notes.", ""),
    ]:
        ws2.append(row)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["C"].width = 60

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["J"].width = 52
    ws.column_dimensions["K"].width = 52

    wb.save(str(path) if _is_path else path)
    return path if _is_path else None


# ── Assessment data extraction ────────────────────────────────────────────────

def extract_assessment_data(models_dict: dict) -> dict[str, dict]:
    """Extract current assessment values from a runtime models dict, keyed by node_id.

    Only elements that carry a 'node_id' field are preserved.
    Returns: {node_id: {support_for, support_against, evidence_for, …}}
    """
    out: dict[str, dict] = {}

    def _grab(d: dict) -> None:
        nid = d.get("node_id")
        if nid:
            out[str(nid)] = {
                "support_for":     float(d.get("support_for", 0.5)),
                "support_against": float(d.get("support_against", 0.1)),
                "evidence_for":    str(d.get("evidence_for", "")),
                "evidence_against": str(d.get("evidence_against", "")),
                "uncertainty_note": str(d.get("uncertainty_note", "")),
                "success_criteria": str(d.get("success_criteria", "")),
            }

    for pdata in models_dict.get("play", {}).values():
        if isinstance(pdata, dict):
            _grab(pdata)
            for sub in pdata.get("sub_elements", []):
                if isinstance(sub, dict):
                    _grab(sub)

    for elems in models_dict.get("conditional", {}).values():
        if isinstance(elems, list):
            for e in elems:
                if isinstance(e, dict):
                    _grab(e)

    return out


# ── Runtime model building ────────────────────────────────────────────────────

def build_runtime_models(
    model: RiskModel,
    assessment_data: "dict[str, dict] | None" = None,
) -> dict:
    """Convert RiskModel + optional assessment_data → {play: …, conditional: …} dict.

    Values in assessment_data override model defaults for matching node_ids.
    Elements always carry a 'node_id' field so they can be matched when the model changes.
    """
    if assessment_data is None:
        assessment_data = {}

    play: dict[str, Any] = {}
    conditional: dict[str, Any] = {}

    for pdef in model.pillars:
        pid = pdef.pillar_id
        pe = pdef.play

        # ── Play ────────────────────────────────────────────────────────────
        if not pe.sub_groups:
            # Single play element — the standard case
            ad = assessment_data.get(pe.node_id, {})
            play[pid] = {
                "label":            pdef.display_name,
                "node_id":          pe.node_id,
                "description":      pe.description or pdef.display_name,
                "considerations":   pe.considerations or "",
                "success_criteria": ad.get("success_criteria", pe.success_criteria),
                "support_for":      ad.get("support_for",     pe.default_s_n),
                "support_against":  ad.get("support_against", pe.default_s_neg),
                "evidence_for":     ad.get("evidence_for",     ""),
                "evidence_against": ad.get("evidence_against", ""),
                "uncertainty_note": ad.get("uncertainty_note", ""),
                "suff_for":   1.0,
                "suff_against": 1.0,
            }
        else:
            # Multi-element play pillar
            ad = assessment_data.get(pe.node_id, {})
            sub_elems: list[dict] = []
            for grp in pe.sub_groups:
                for elem in grp.elements:
                    ead = assessment_data.get(elem.node_id, {})
                    sub_elems.append({
                        "label":            grp.group_label,
                        "node_id":          elem.node_id,
                        "success_criteria": ead.get("success_criteria", elem.success_criteria),
                        "support_for":      ead.get("support_for",     elem.default_s_n),
                        "support_against":  ead.get("support_against", elem.default_s_neg),
                        "evidence_for":     ead.get("evidence_for",     ""),
                        "evidence_against": ead.get("evidence_against", ""),
                        "uncertainty_note": ead.get("uncertainty_note", ""),
                        "suff_for":   1.0,
                        "suff_against": 1.0,
                    })
            play[pid] = {
                "label":            pdef.display_name,
                "node_id":          pe.node_id,
                "description":      pe.description or pdef.display_name,
                "considerations":   pe.considerations or "",
                "success_criteria": ad.get("success_criteria", pe.success_criteria),
                "support_for":      ad.get("support_for",     pe.default_s_n),
                "support_against":  ad.get("support_against", pe.default_s_neg),
                "evidence_for":     ad.get("evidence_for",     ""),
                "evidence_against": ad.get("evidence_against", ""),
                "uncertainty_note": ad.get("uncertainty_note", ""),
                "suff_for":   1.0,
                "suff_against": 1.0,
                "sub_elements":    sub_elems,
                "aggregation_rule": pe.aggregation_rule,
            }

        # ── Conditional ─────────────────────────────────────────────────────
        cond_elems: list[dict] = []
        for grp in pdef.conditional_groups:
            for elem in grp.elements:
                ead = assessment_data.get(elem.node_id, {})
                cond_elems.append({
                    "label":            grp.group_label,
                    "node_id":          elem.node_id,
                    "success_criteria": ead.get("success_criteria", elem.success_criteria),
                    "support_for":      ead.get("support_for",     elem.default_s_n),
                    "support_against":  ead.get("support_against", elem.default_s_neg),
                    "evidence_for":     ead.get("evidence_for",     ""),
                    "evidence_against": ead.get("evidence_against", ""),
                    "uncertainty_note": ead.get("uncertainty_note", ""),
                    "suff_for":   1.0,
                    "suff_against": 1.0,
                    "_group_operator": grp.aggregation_rule,
                })
        conditional[pid] = cond_elems

    return {"play": play, "conditional": conditional}


# ── Operator initializer ──────────────────────────────────────────────────────

def initialize_operators_from_model(model: RiskModel, ss: dict) -> None:
    """Pre-fill ESL operator session-state keys from model definition.

    Call this after build_runtime_models() to ensure the Conditional tab uses
    the operators specified in the model template.

    ss: the session_state dict (or any mapping) to update in-place.
    """
    def _nearest(op: str) -> str:
        if op in _ESL_PIPELINE_MODE_OPTIONS:
            return op
        lop = op.lower()
        if "all" in lop or "min" in lop:
            return "ESL-ALL (min/min)"
        if "any" in lop or "max" in lop:
            return "ESL-ANY (max/max)"
        if "ipt" in lop or "suff" in lop:
            return "ESL-IPT (sufficiency/dependency)"
        if "product" in lop or "π" in lop:
            return "Product (Π)"
        if "mean" in lop or "avg" in lop:
            return "Mean"
        return "ESL-ALL (min/min)"

    for pdef in model.pillars:
        pid = pdef.pillar_id
        ss[SK.esl_mode(pid)] = _nearest(pdef.conditional_operator)
        for grp in pdef.conditional_groups:
            ss[SK.esl_group_mode(pid, grp.group_label)] = _nearest(grp.aggregation_rule)


def initialize_classic_operators_from_model(model: RiskModel, ss: dict) -> None:
    """Pre-fill Classic POS operator session-state keys from model definition.

    Call this alongside initialize_operators_from_model() after loading/applying a model.
    Keys written: classic_mode_cond_{pillar_id} and classic_mode_group_cond_{pillar_id}_{group_label}.

    ss: the session_state dict (or any mapping) to update in-place.
    """
    for pdef in model.pillars:
        pid = pdef.pillar_id
        ss[SK.classic_mode(pid)] = pdef.classic_pos_operator
        for grp in pdef.conditional_groups:
            ss[SK.classic_group_mode(pid, grp.group_label)] = grp.classic_pos_aggregation_rule


# ── Default model bootstrap ────────────────────────────────────────────────────

def ensure_default_model() -> RiskModel:
    """Load or create the default model from data/models/default.json.

    Priority:
      1. Load existing data/models/default.json
      2. Build from risk_model_nodes.xlsx (project root)
      3. Raise FileNotFoundError
    """
    default_path = MODELS_DIR / "default.json"

    if default_path.exists():
        return RiskModel.load(default_path)

    xlsx_candidates = [
        Path("risk_model_nodes.xlsx"),
        Path(__file__).parent.parent / "risk_model_nodes.xlsx",
    ]
    for xlsx_path in xlsx_candidates:
        if xlsx_path.exists():
            model = from_xlsx(xlsx_path, "default", "Standard 4-Pillar ESL Model")
            MODELS_DIR.mkdir(parents=True, exist_ok=True)
            model.save(default_path)
            return model

    raise FileNotFoundError(
        "Cannot find default.json or risk_model_nodes.xlsx. "
        "Place risk_model_nodes.xlsx in the project root, or use the Model Builder to create a model."
    )


# ── Private helpers ────────────────────────────────────────────────────────────

def _float_safe(v: Any, default: float = 0.5) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _slug(s: str, maxlen: int = 20) -> str:
    """Convert a string to a compact uppercase slug for node_id generation."""
    import re
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").upper()
    return s[:maxlen]
